"""Module for cleaning document metadata."""

import io
from pathlib import Path
from typing import Any, Optional
from dataclasses import dataclass, field

# Document processing
from docx import Document
import openpyxl

# PDF handling
try:
    import pikepdf
    HAS_PIKEPDF = True
except ImportError:
    HAS_PIKEPDF = False

# Image metadata
from PIL import Image
from PIL.ExifTags import TAGS


@dataclass
class CleanedMetadata:
    """Information about cleaned metadata."""
    original_fields: dict = field(default_factory=dict)
    removed_fields: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


class MetadataCleaner:
    """Clean metadata from various document types."""
    
    def clean_docx_metadata(self, file_path: Path, output_path: Path) -> CleanedMetadata:
        """
        Clean metadata from a Word document.
        
        Args:
            file_path: Path to original document
            output_path: Path to save cleaned document
            
        Returns:
            CleanedMetadata with information about what was cleaned
        """
        result = CleanedMetadata()
        
        doc = Document(file_path)
        core_props = doc.core_properties
        
        # Store original metadata
        result.original_fields = {
            "author": core_props.author,
            "title": core_props.title,
            "subject": core_props.subject,
            "keywords": core_props.keywords,
            "comments": core_props.comments,
            "last_modified_by": core_props.last_modified_by,
            "revision": core_props.revision,
            "category": core_props.category,
        }
        
        # Clean metadata
        if core_props.author:
            result.removed_fields.append("author")
            core_props.author = ""
        
        if core_props.title:
            result.removed_fields.append("title")
            core_props.title = ""
        
        if core_props.subject:
            result.removed_fields.append("subject")
            core_props.subject = ""
        
        if core_props.keywords:
            result.removed_fields.append("keywords")
            core_props.keywords = ""
        
        if core_props.comments:
            result.removed_fields.append("comments")
            core_props.comments = ""
        
        if core_props.last_modified_by:
            result.removed_fields.append("last_modified_by")
            core_props.last_modified_by = ""
        
        if core_props.category:
            result.removed_fields.append("category")
            core_props.category = ""
        
        # Note: revision can't be set to 0, leave as is
        
        # Save cleaned document
        doc.save(output_path)
        
        return result
    
    def clean_xlsx_metadata(self, file_path: Path, output_path: Path) -> CleanedMetadata:
        """
        Clean metadata from an Excel workbook.
        
        Args:
            file_path: Path to original workbook
            output_path: Path to save cleaned workbook
            
        Returns:
            CleanedMetadata with information about what was cleaned
        """
        result = CleanedMetadata()
        
        wb = openpyxl.load_workbook(file_path)
        props = wb.properties
        
        # Store original metadata
        result.original_fields = {
            "creator": props.creator,
            "title": props.title,
            "subject": props.subject,
            "keywords": props.keywords,
            "description": props.description,
            "lastModifiedBy": props.lastModifiedBy,
            "category": props.category,
        }
        
        # Clean metadata
        if props.creator:
            result.removed_fields.append("creator")
            props.creator = ""
        
        if props.title:
            result.removed_fields.append("title")
            props.title = ""
        
        if props.subject:
            result.removed_fields.append("subject")
            props.subject = ""
        
        if props.keywords:
            result.removed_fields.append("keywords")
            props.keywords = ""
        
        if props.description:
            result.removed_fields.append("description")
            props.description = ""
        
        if props.lastModifiedBy:
            result.removed_fields.append("lastModifiedBy")
            props.lastModifiedBy = ""
        
        if props.category:
            result.removed_fields.append("category")
            props.category = ""
        
        # Remove comments from cells
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.comment:
                        cell.comment = None
                        result.removed_fields.append(f"comment_{sheet_name}_{cell.coordinate}")
        
        # Save cleaned workbook
        wb.save(output_path)
        wb.close()
        
        return result
    
    def clean_pdf_metadata(self, file_path: Path, output_path: Path) -> CleanedMetadata:
        """
        Clean metadata from a PDF document.
        
        Args:
            file_path: Path to original PDF
            output_path: Path to save cleaned PDF
            
        Returns:
            CleanedMetadata with information about what was cleaned
        """
        result = CleanedMetadata()
        
        if not HAS_PIKEPDF:
            result.warnings.append("pikepdf not installed - PDF metadata cleaning limited")
            import shutil
            shutil.copy(file_path, output_path)
            return result
        
        try:
            with pikepdf.open(file_path) as pdf:
                # Store original metadata
                if pdf.docinfo:
                    for key in list(pdf.docinfo.keys()):
                        try:
                            result.original_fields[str(key)] = str(pdf.docinfo[key])
                        except Exception:
                            pass
                
                # Remove docinfo keys one by one (clear() doesn't work reliably)
                if pdf.docinfo:
                    keys_to_remove = list(pdf.docinfo.keys())
                    for key in keys_to_remove:
                        try:
                            del pdf.docinfo[key]
                            result.removed_fields.append(str(key))
                        except Exception:
                            pass
                
                # Try to remove XMP metadata
                try:
                    if "/Metadata" in pdf.Root:
                        del pdf.Root["/Metadata"]
                        result.removed_fields.append("XMP_metadata")
                except Exception:
                    pass
                
                # Remove any annotations that might contain identifying info
                for page in pdf.pages:
                    if "/Annots" in page:
                        try:
                            del page["/Annots"]
                            result.removed_fields.append("annotations")
                        except Exception:
                            pass
                
                # Save cleaned PDF
                pdf.save(output_path)
                
        except Exception as e:
            result.warnings.append(f"PDF metadata cleaning error: {str(e)}")
            import shutil
            shutil.copy(file_path, output_path)
        
        return result
    
    def clean_image_metadata(self, image_data: bytes) -> tuple[bytes, CleanedMetadata]:
        """
        Clean EXIF and other metadata from an image.
        
        Args:
            image_data: Image as bytes
            
        Returns:
            Tuple of (cleaned image bytes, CleanedMetadata)
        """
        result = CleanedMetadata()
        
        try:
            # Open image
            img = Image.open(io.BytesIO(image_data))
            
            # Extract original EXIF data
            exif_data = img.getexif()
            if exif_data:
                for tag_id, value in exif_data.items():
                    tag_name = TAGS.get(tag_id, str(tag_id))
                    result.original_fields[tag_name] = str(value)
                    result.removed_fields.append(tag_name)
            
            # Create new image without metadata
            output = io.BytesIO()
            
            # Convert to RGB if necessary (removes alpha channel issues)
            if img.mode in ("RGBA", "P"):
                img = img.convert("RGB")
            
            # Save without metadata
            img.save(output, format="PNG", optimize=True)
            
            return output.getvalue(), result
            
        except Exception as e:
            result.warnings.append(f"Image metadata cleaning error: {str(e)}")
            return image_data, result
    
    def clean_document(
        self, 
        file_path: Path, 
        output_path: Path, 
        file_type: str
    ) -> CleanedMetadata:
        """
        Clean metadata from a document based on its type.
        
        Args:
            file_path: Path to original document
            output_path: Path to save cleaned document
            file_type: Type of document (docx, xlsx, pdf)
            
        Returns:
            CleanedMetadata with information about what was cleaned
        """
        if file_type == "docx":
            return self.clean_docx_metadata(file_path, output_path)
        elif file_type in ["xlsx", "xls"]:
            return self.clean_xlsx_metadata(file_path, output_path)
        elif file_type == "pdf":
            return self.clean_pdf_metadata(file_path, output_path)
        else:
            result = CleanedMetadata()
            result.warnings.append(f"Unsupported file type for metadata cleaning: {file_type}")
            return result


